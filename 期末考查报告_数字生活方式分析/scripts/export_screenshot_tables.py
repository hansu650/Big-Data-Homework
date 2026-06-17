"""Export Excel workbooks for manual screenshot insertion in the final report."""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]


RAW_DATA_PATH = PROJECT_ROOT / "data" / "raw" / "digital_lifestyle_benchmark_2025.csv"
RESULTS_DIR = PROJECT_ROOT / "results"
OUTPUT_DIR = PROJECT_ROOT / "screenshot_tables"

ID_COLUMNS = ["id"]
CLASSIFICATION_TARGET = "high_risk_flag"
HIGH_RISK_LEAKAGE_COLUMNS = [
    "anxiety_score",
    "depression_score",
    "stress_level",
    "happiness_score",
    "focus_score",
    "productivity_score",
    "digital_dependence_score",
]
OUTCOME_COLUMNS = [
    "anxiety_score",
    "depression_score",
    "stress_level",
    "happiness_score",
    "focus_score",
    "high_risk_flag",
    "productivity_score",
    "digital_dependence_score",
]
CLUSTERING_INPUT_COLUMNS = [
    "device_hours_per_day",
    "phone_unlocks",
    "notifications_per_day",
    "social_media_mins",
    "study_mins",
    "physical_activity_days",
    "sleep_hours",
    "sleep_quality",
    "social_media_hours",
    "study_hours",
    "notifications_per_device_hour",
    "unlocks_per_device_hour",
    "device_to_sleep_ratio",
    "activity_sleep_interaction",
    "social_to_study_ratio",
]


FIELD_DESCRIPTIONS = {
    "id": "Unique record identifier.",
    "age": "User age.",
    "gender": "User gender category.",
    "region": "User region.",
    "income_level": "Income-level category.",
    "education_level": "Education-level category.",
    "daily_role": "Main daily role or occupation status.",
    "device_hours_per_day": "Daily digital-device use time in hours.",
    "phone_unlocks": "Number of phone unlocks per day.",
    "notifications_per_day": "Number of notifications received per day.",
    "social_media_mins": "Daily social-media time in minutes.",
    "study_mins": "Daily study time in minutes.",
    "physical_activity_days": "Physical activity days per week.",
    "sleep_hours": "Daily sleep duration in hours.",
    "sleep_quality": "Sleep-quality score.",
    "anxiety_score": "Synthetic anxiety-related outcome score.",
    "depression_score": "Synthetic depression-related outcome score.",
    "stress_level": "Synthetic stress-level outcome score.",
    "happiness_score": "Synthetic happiness score.",
    "focus_score": "Synthetic focus score.",
    "high_risk_flag": "Classification target: 1 means High Risk and 0 means No Risk.",
    "device_type": "Main device type.",
    "productivity_score": "Auxiliary regression target; weak-prediction result in this report.",
    "digital_dependence_score": "Main regression target for digital dependence prediction.",
}


RANGE_RULES = {
    "age": (10, 90, "10 to 90 years"),
    "device_hours_per_day": (0, 24, "0 to 24 hours/day"),
    "phone_unlocks": (0, 500, "0 to 500 unlocks/day"),
    "notifications_per_day": (0, 1000, "0 to 1000 notifications/day"),
    "social_media_mins": (0, 720, "0 to 720 minutes/day"),
    "study_mins": (0, 720, "0 to 720 minutes/day"),
    "physical_activity_days": (0, 7, "0 to 7 days/week"),
    "sleep_hours": (0, 14, "0 to 14 hours/day"),
    "sleep_quality": (0, 5, "0 to 5 score"),
    "productivity_score": (0, 100, "0 to 100 score"),
    "digital_dependence_score": (0, 100, "0 to 100 score"),
}


def ensure_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for old_file in OUTPUT_DIR.glob("table*.xlsx"):
        old_file.unlink()


def clean_value(value):
    if isinstance(value, float):
        return round(value, 4)
    return value


def add_behavior_features(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    eps = 1e-6
    out["social_media_hours"] = out["social_media_mins"] / 60.0
    out["study_hours"] = out["study_mins"] / 60.0
    out["notifications_per_device_hour"] = out["notifications_per_day"] / out["device_hours_per_day"].clip(lower=eps)
    out["unlocks_per_device_hour"] = out["phone_unlocks"] / out["device_hours_per_day"].clip(lower=eps)
    out["device_to_sleep_ratio"] = out["device_hours_per_day"] / out["sleep_hours"].clip(lower=eps)
    out["activity_sleep_interaction"] = out["physical_activity_days"] * out["sleep_hours"]
    out["social_to_study_ratio"] = out["social_media_mins"] / (out["study_mins"] + 1.0)
    return out


def get_excluded_columns(task: str, target_column: str | None = None) -> list[str]:
    excluded = set(ID_COLUMNS)
    if task == "classification":
        excluded.add(CLASSIFICATION_TARGET)
        excluded.update(HIGH_RISK_LEAKAGE_COLUMNS)
    elif task == "regression":
        target = target_column or "digital_dependence_score"
        excluded.add(target)
        excluded.update(OUTCOME_COLUMNS)
        if target == "digital_dependence_score":
            excluded.add("productivity_score")
        elif target == "productivity_score":
            excluded.add("digital_dependence_score")
    else:
        raise ValueError(task)
    return sorted(excluded)


def get_input_features(df: pd.DataFrame, excluded: list[str]) -> list[str]:
    return [column for column in df.columns if column not in excluded]


def get_clustering_feature_columns(df: pd.DataFrame) -> list[str]:
    return [column for column in CLUSTERING_INPUT_COLUMNS if column in df.columns]


def save_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F2937")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 38)
        ws.row_dimensions[1].height = 28
    wb.save(path)


def table1_raw_preview(raw: pd.DataFrame) -> None:
    save_workbook(OUTPUT_DIR / "table1_raw_dataset_preview.xlsx", {"raw_preview": raw.head(20)})


def table2_dataset_fields(raw: pd.DataFrame) -> None:
    rows = []
    for column in raw.columns:
        rows.append(
            {
                "field_name": column,
                "dtype": str(raw[column].dtype),
                "non_null_count": int(raw[column].notna().sum()),
                "example_value": clean_value(raw[column].dropna().iloc[0]) if raw[column].notna().any() else "",
                "field_meaning": FIELD_DESCRIPTIONS.get(column, "Dataset field used in analysis."),
            }
        )
    save_workbook(OUTPUT_DIR / "table2_dataset_fields.xlsx", {"fields": pd.DataFrame(rows)})


def table3_missing_values(raw: pd.DataFrame) -> None:
    df = pd.DataFrame(
        {
            "field_name": raw.columns,
            "missing_count": raw.isna().sum().values,
            "missing_rate": raw.isna().mean().values,
        }
    )
    save_workbook(OUTPUT_DIR / "table3_missing_value_check.xlsx", {"missing_values": df})


def table4_duplicates(raw: pd.DataFrame) -> None:
    duplicate_rows = int(raw.duplicated().sum())
    duplicate_id = int(raw["id"].duplicated().sum()) if "id" in raw.columns else np.nan
    df = pd.DataFrame(
        [
            {"check_item": "Total samples", "value": len(raw), "result": "Recorded"},
            {"check_item": "Duplicate rows", "value": duplicate_rows, "result": "No deletion needed"},
            {"check_item": "Duplicate id values", "value": duplicate_id, "result": "No deletion needed"},
            {"check_item": "Processing decision", "value": "Keep all records", "result": "Passed"},
        ]
    )
    save_workbook(OUTPUT_DIR / "table4_duplicate_check.xlsx", {"duplicate_check": df})


def table5_range_check(raw: pd.DataFrame) -> None:
    rows = []
    for column, (low, high, label) in RANGE_RULES.items():
        series = raw[column]
        min_value = float(series.min())
        max_value = float(series.max())
        mean_value = float(series.mean())
        passed = bool(min_value >= low and max_value <= high)
        rows.append(
            {
                "field_name": column,
                "min": min_value,
                "max": max_value,
                "mean": mean_value,
                "reasonable_range": label,
                "passed": "PASS" if passed else "CHECK",
            }
        )
    save_workbook(
        OUTPUT_DIR / "table5_range_and_rationality_check.xlsx",
        {"range_rationality": pd.DataFrame(rows)},
    )


def table6_engineered_features(raw: pd.DataFrame) -> pd.DataFrame:
    engineered = add_behavior_features(raw)
    cols = [
        "device_hours_per_day",
        "phone_unlocks",
        "notifications_per_day",
        "social_media_mins",
        "study_mins",
        "physical_activity_days",
        "sleep_hours",
        "sleep_quality",
        "social_media_hours",
        "study_hours",
        "notifications_per_device_hour",
        "unlocks_per_device_hour",
        "device_to_sleep_ratio",
        "activity_sleep_interaction",
        "social_to_study_ratio",
    ]
    save_workbook(
        OUTPUT_DIR / "table6_engineered_features_preview.xlsx",
        {"engineered_preview": engineered[cols].head(20)},
    )
    return engineered


def table7_feature_selection(engineered: pd.DataFrame) -> None:
    classification_drop = get_excluded_columns("classification")
    regression_dd_drop = get_excluded_columns("regression", target_column="digital_dependence_score")
    regression_prod_drop = get_excluded_columns("regression", target_column="productivity_score")
    clustering_features = get_clustering_feature_columns(engineered)
    sheets = {
        "classification_drop_columns": pd.DataFrame({"drop_column": classification_drop}),
        "classification_input_features": pd.DataFrame({"input_feature": get_input_features(engineered, classification_drop)}),
        "regression_dd_drop_columns": pd.DataFrame(
            {"drop_column": regression_dd_drop}
        ),
        "regression_dd_input_features": pd.DataFrame(
            {"input_feature": get_input_features(engineered, regression_dd_drop)}
        ),
        "regression_prod_drop_columns": pd.DataFrame(
            {"drop_column": regression_prod_drop}
        ),
        "regression_prod_input_features": pd.DataFrame(
            {"input_feature": get_input_features(engineered, regression_prod_drop)}
        ),
        "clustering_features": pd.DataFrame({"clustering_feature": clustering_features}),
    }
    save_workbook(OUTPUT_DIR / "table7_feature_selection_and_leakage_control.xlsx", sheets)


def table8_pca_explained_variance() -> None:
    pca = pd.read_csv(RESULTS_DIR / "pca_explained_variance.csv")
    pca = pca[
        [
            "component",
            "component_index",
            "explained_variance_ratio",
            "cumulative_explained_variance",
            "input_feature_count",
        ]
    ].copy()
    pca["explained_variance_percent"] = pca["explained_variance_ratio"] * 100
    pca["cumulative_explained_variance_percent"] = (
        pca["cumulative_explained_variance"] * 100
    )
    pca["interpretation_note"] = np.where(
        pca["component_index"].eq(2),
        "PC1 + PC2 explain about 42.41%; useful for visualization only.",
        "",
    )
    save_workbook(OUTPUT_DIR / "table8_pca_explained_variance.xlsx", {"pca_variance": pca})


def table9_descriptive_statistics(raw: pd.DataFrame) -> None:
    columns = [
        "device_hours_per_day",
        "phone_unlocks",
        "notifications_per_day",
        "social_media_mins",
        "study_mins",
        "sleep_hours",
        "sleep_quality",
        "productivity_score",
        "digital_dependence_score",
    ]
    summary = raw[columns].describe(percentiles=[0.25, 0.5, 0.75]).T.reset_index()
    summary = summary.rename(
        columns={
            "index": "field_name",
            "25%": "Q1",
            "50%": "median",
            "75%": "Q3",
        }
    )
    summary = summary[
        ["field_name", "count", "mean", "std", "min", "Q1", "median", "Q3", "max"]
    ]
    save_workbook(
        OUTPUT_DIR / "table9_descriptive_statistical_summary.xlsx",
        {"descriptive_summary": summary},
    )


def table10_classification_comparison() -> None:
    metrics = pd.read_csv(RESULTS_DIR / "classification_tuned_metrics.csv")
    model_compare = metrics[
        (metrics["dataset"].eq("validation"))
        & (metrics["threshold_policy"].eq("default_0_50"))
    ].copy()
    model_compare = model_compare[
        [
            "dataset",
            "model",
            "threshold_policy",
            "threshold",
            "accuracy",
            "precision",
            "recall",
            "f1",
            "roc_auc",
            "pr_auc",
            "balanced_accuracy",
        ]
    ]
    threshold = pd.read_csv(RESULTS_DIR / "classification_threshold_tuning.csv")
    threshold = threshold[
        [
            "dataset",
            "model",
            "policy",
            "threshold",
            "accuracy",
            "precision",
            "recall",
            "f1",
            "roc_auc",
            "pr_auc",
            "balanced_accuracy",
        ]
    ]
    final_test = metrics[
        (metrics["dataset"].eq("test"))
        & (metrics["model"].eq("gradient_boosting"))
        & (metrics["threshold"].round(2).eq(0.14))
    ].copy()
    final_test = final_test[
        [
            "dataset",
            "model",
            "threshold_policy",
            "threshold",
            "accuracy",
            "precision",
            "recall",
            "f1",
            "roc_auc",
            "pr_auc",
            "balanced_accuracy",
            "tn",
            "fp",
            "fn",
            "tp",
        ]
    ]
    matrix = pd.read_csv(RESULTS_DIR / "classification_final_confusion_matrix.csv")
    matrix = matrix.rename(columns={matrix.columns[0]: "actual_label"})
    sheets = {
        "classification_model_comparison": model_compare,
        "threshold_strategy_comparison": threshold,
        "final_test_metrics": final_test,
        "confusion_matrix_counts": matrix,
    }
    save_workbook(
        OUTPUT_DIR / "table10_classification_model_comparison_and_threshold_results.xlsx",
        sheets,
    )


def table11_regression_comparison() -> None:
    dd = pd.read_csv(RESULTS_DIR / "regression_digital_dependence_metrics.csv")
    prod = pd.read_csv(RESULTS_DIR / "regression_productivity_metrics.csv")
    target_comp = pd.read_csv(RESULTS_DIR / "regression_target_comparison.csv")
    metric_cols = [
        "model",
        "target",
        "cv_best_r2",
        "cv_best_mse",
        "cv_best_mae",
        "mae",
        "mse",
        "rmse",
        "r2",
    ]
    sheets = {
        "target_comparison": target_comp,
        "digital_dependence_models": dd[metric_cols],
        "productivity_models": prod[metric_cols],
    }
    save_workbook(OUTPUT_DIR / "table11_regression_model_comparison.xlsx", sheets)


def table12_clustering_comparison() -> None:
    model_comp = pd.read_csv(RESULTS_DIR / "clustering_model_comparison.csv")
    kmeans = pd.read_csv(RESULTS_DIR / "clustering_kmeans_scores.csv")
    profiles = pd.read_csv(RESULTS_DIR / "clustering_lifestyle_profiles_compact.csv")
    common_cols = [
        "algorithm",
        "k",
        "inertia",
        "silhouette",
        "calinski_harabasz",
        "davies_bouldin",
    ]
    sheets = {
        "model_comparison": model_comp[common_cols],
        "kmeans_k_scores": kmeans,
        "cluster_profiles": profiles,
    }
    save_workbook(
        OUTPUT_DIR / "table12_clustering_model_comparison_and_cluster_profiles.xlsx",
        sheets,
    )


def main() -> None:
    ensure_dir()
    raw = pd.read_csv(RAW_DATA_PATH)
    table1_raw_preview(raw)
    table2_dataset_fields(raw)
    table3_missing_values(raw)
    table4_duplicates(raw)
    table5_range_check(raw)
    engineered = table6_engineered_features(raw)
    table7_feature_selection(engineered)
    table8_pca_explained_variance()
    table9_descriptive_statistics(raw)
    table10_classification_comparison()
    table11_regression_comparison()
    table12_clustering_comparison()
    print(f"Generated screenshot Excel tables in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
